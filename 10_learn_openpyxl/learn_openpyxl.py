"""
    Author: shikechen
    Function: Learn openpyxl lib, reference http://zetcode.com/articles/openpyxl/
    Version: 1.0
    Date: 2019/3/5
"""
from openpyxl import Workbook
import time
import openpyxl


def main():
    # book = Workbook()
    # sheet = book.active

    # Example 1
    # sheet['A1'] = 56
    # sheet['A2'] = 43
    #
    # now = time.strftime('%x')
    # sheet['A3'] = now
    #
    # book.save('sample.xlsx')

    # Example 2
    # sheet['A1'] = 1
    # sheet.cell(row=2, column=2).value = 2
    #
    # book.save('write2cell.xlsx')

    # Example 3
    # rows = (
    #     (88, 46, 57),
    #     (89, 38, 12),
    #     (23, 59, 78),
    #     (56, 21, 98),
    #     (24, 18, 43),
    #     (34, 15, 67)
    # )
    #
    # for row in rows:
    #     sheet.append(row)
    #
    # book.save('appending.xlsx')

    # Example 4
    # book = openpyxl.load_workbook('sample.xlsx')
    # sheet = book.active
    #
    # a1 = sheet['A1']
    # a2 = sheet['A2']
    # a3 = sheet.cell(row=3, column=1)
    #
    # print(a1.value)
    # print(a2.value)
    # print(a3.value)

    # Example 5
    # book = openpyxl.load_workbook('items.xlsx')
    # sheet = book.active
    # cells = sheet['A1': 'B6']
    #
    # for c1, c2 in cells:
    #     print('{0:8} {1:8}'.format(c1.value, c2.value))

    # Example 6
    book = Workbook()
    sheet = book.active

    rows = (
        (88, 46, 57),
        (89, 38, 12),
        (23, 59, 78),
        (56, 21, 98),
        (24, 18, 43),
        (34, 15, 67)
    )

    for row in rows:
        sheet.append(row)

    # for row in sheet.iter_rows(min_row=1, min_col=1, max_row=6, max_col=3):
    #     for cell in row:
    #         print(cell.value, end=' ')
    #     print()
    #
    # book.save('iterbyrows.xlsx')

    for row in sheet.iter_cols(min_row=1, min_col=1, max_row=6, max_col=3):
        for cell in row:
            print(cell.value, end=' ')
        print()

    book.save('iterbycols.xlsx')


if __name__ == '__main__':
    main()
