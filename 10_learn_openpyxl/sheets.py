"""
    Author: shikechen
    Function: Learn multi sheets operation
    Version: 1.0
    Date: 2019/3/5
"""
import openpyxl


def main():
    # book = openpyxl.load_workbook('2019.xlsx')
    # print(book.sheetnames)
    #
    # active_sheet = book.active
    # print(type(active_sheet))
    #
    # sheet = book['March']
    # print(sheet.title)

    # book = openpyxl.load_workbook('2019.xlsx')
    # book.create_sheet('April')
    #
    # print(book.sheetnames)
    #
    # sheet1 = book['January']
    # book.remove(sheet1)
    #
    # print(book.sheetnames)
    #
    # book.create_sheet('January', 0)
    # print(book.sheetnames)
    #
    # book.save('2019_2.xlsx')

    # change the background color of a worksheet
    book = openpyxl.load_workbook('2019.xlsx')
    sheet = book['March']
    sheet.sheet_properties.tabColor = '0072BA'

    book.save('2019_3.xlsx')


if __name__ == '__main__':
    main()
