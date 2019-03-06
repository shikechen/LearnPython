"""
    Author: shikechen
    Function: Learn dimensions
    Version: 1.0
    Date: 2019/3/5
"""
from openpyxl import Workbook


def main():
    book = Workbook()
    sheet = book.active

    sheet['A3'] = 39
    sheet['B3'] = 19

    rows = [
        (88, 46),
        (89, 38),
        (23, 59),
        (56, 21),
        (24, 18),
        (34, 15)
    ]

    for row in rows:
        sheet.append(row)

    print(sheet.dimensions)
    print('Minimum row: {0}'.format(sheet.min_row))
    print('Maximum row: {0}'.format(sheet.max_row))
    print('Minimum column: {0}'.format(sheet.min_column))
    print('Maximum column: {0}'.format(sheet.max_column))

    for c1, c2 in sheet[sheet.dimensions]:
        print(c1.value, c2.value)

    book.save('dimensions.xlsx')


if __name__ == '__main__':
    main()
