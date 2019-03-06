"""
    Author: shikechen
    Function: Learn merging cells operation
    Version: 1.0
    Date: 2019/3/5
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment


def main():
    book = Workbook()
    sheet = book.active

    sheet.merge_cells('A1:B2')

    cell = sheet.cell(row=1, column=1)
    cell.value = 'Sunny day'
    cell.alignment = Alignment(horizontal='center', vertical='center')

    book.save('merging.xlsx')


if __name__ == '__main__':
    main()
