"""
    Author: shikechen
    Function: Company staff meal report intelligent statistical system
    Version: 1.2
    Date: 2019/3/11
"""
import openpyxl
from openpyxl.utils import get_column_letter


def main():
    excel_name = '2019_2.xlsx'
    book = openpyxl.load_workbook(excel_name)
    sheet_name_list = book.sheetnames
    print(sheet_name_list)

    for sheet_name in sheet_name_list:
        print('sheet name: ', sheet_name)
        if sheet_name.__contains__('-'):
            sheet = book[sheet_name]
            print('sheet title: ', sheet.title)

            rows = sheet.rows

            sheet_max_row = sheet.max_row
            print('sheet_max_row', sheet_max_row)
            sheet_max_column = sheet.max_column
            print('sheet_max_column', sheet_max_column)

            need_count_columns = []
            row_index = 0
            column_write_index = sheet_max_column + 1
            for row in rows:
                row_sum = 0
                row_index += 1

                # the first row is title
                if row_index > 1:
                    for cell in row:
                        cell_value = cell.value
                        # print(cell_value)

                        cell_sum = 0

                        if isinstance(cell_value, str):
                            if '早餐' in cell_value:
                                cell_sum += 5

                            if '午餐' in cell_value:
                                cell_sum += 15

                            if '晚餐' in cell_value:
                                cell_sum += 10

                        row_sum += cell_sum

                    sheet.cell(row=row_index, column=column_write_index).value = row_sum

                else:
                    for cell in row:
                        cell_value = cell.value
                        if '月' in cell_value:
                            print('This column need to count ', cell.column)
                            print('This column need to count coordinate ', cell.coordinate)
                            need_count_columns.append(cell.column)

            print('need_count_columns ', need_count_columns)
            print('sheet_max_column', sheet_max_column)

            row_write_index = sheet_max_row + 1

            for count_column in need_count_columns:
                breakfast_count = 0
                lunch_count = 0
                dinner_count = 0
                for cell in sheet[get_column_letter(count_column)]:
                    cell_value = cell.value
                    print('cell value ', cell_value)

                    if isinstance(cell_value, str):
                        if '早餐' in cell_value:
                            breakfast_count += 1

                        if '午餐' in cell_value:
                            lunch_count += 1

                        if '晚餐' in cell_value:
                            dinner_count += 1

                result_str = '早餐:{}, 午餐:{}, 晚餐:{}'.format(breakfast_count, lunch_count, dinner_count)
                sheet.cell(row=row_write_index, column=count_column).value = result_str

    book.save(excel_name)
    book.close()


if __name__ == '__main__':
    main()
